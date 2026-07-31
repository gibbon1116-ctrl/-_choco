from openpyxl import load_workbook

from src import db
from src.excel_io import export_schedule, load_sample_data
from src.restaurant import SKILL_LEVEL_LABELS, employee_has_role
from src.solver import generate_schedule


def test_restaurant_sample_schedule_and_export_privacy(tmp_path):
    database = tmp_path / "restaurant.sqlite"
    loaded = load_sample_data("2026-08", database, tmp_path / "samples")
    assert loaded == {
        "employees": 12, "requirements": 93, "requests": 6, "relations": 5,
        "campaigns": 2, "role_requirements": 93, "events": 3,
    }

    employees = {e["employee_id"]: e for e in db.fetch_all("employees", database)}
    assert sum(e["english_level"] != "none" for e in employees.values()) == 4
    assert sum(e["new_product_skill"] == 3 for e in employees.values()) == 3
    assert sum(bool(e["can_cashier"]) for e in employees.values()) == 8
    assert sum(bool(e["is_new_staff"]) for e in employees.values()) == 2

    result = generate_schedule("2026-08", 30, database)
    assert result["status"] == "success"
    assert result["restaurant_warnings"] == []

    assignments = {(a["employee_id"], a["date"]): a["shift_code"] for a in result["assignments"]}
    required_days = {r["date"] for r in db.fetch_all(
        "requirements", database, where="target_month=?", params=("2026-08",)) if r["required_count"] > 0}
    for day in required_days:
        assert any(assignments[eid, day] != "O" and employees[eid]["english_level"] != "none"
                   for eid in employees)
        assert not (assignments["E009", day] == assignments["E010", day] != "O")

    schedule_id = result["summary"]["schedule_id"]
    normal = export_schedule(schedule_id, tmp_path, database, admin_export=False)
    admin = export_schedule(schedule_id, tmp_path, database, admin_export=True)
    normal_sheets = load_workbook(normal, read_only=True).sheetnames
    admin_sheets = load_workbook(admin, read_only=True).sheetnames
    assert "飲食店条件確認" in normal_sheets
    assert "スタッフスキル一覧" in normal_sheets
    assert "新商品イベント一覧" in normal_sheets
    assert "相性条件確認" not in normal_sheets
    assert "相性条件確認" in admin_sheets


def test_store_settings_link_english_level_and_required_counts(tmp_path):
    database = tmp_path / "store_settings.sqlite"
    db.init_db(database)
    settings = db.get_store_settings(database)
    settings.update({"required_english_level": "conversational",
                     "required_english_count": 2,
                     "required_new_product_count": 3,
                     "required_allergy_count": 2})
    db.save_store_settings(settings, database)

    saved = db.get_store_settings(database)
    assert saved["required_english_level"] == "conversational"
    assert saved["required_english_count"] == 2
    assert saved["required_new_product_count"] == 3
    assert saved["required_allergy_count"] == 2


def test_english_requirement_uses_minimum_staff_level_and_skill_labels():
    employee = {"english_level": "conversational"}
    assert employee_has_role(employee, "english_support", skill_level=2)
    assert not employee_has_role(employee, "english_support", skill_level=3)
    assert SKILL_LEVEL_LABELS == {0: "未経験", 1: "補助できる", 2: "一人で対応できる", 3: "指導できる"}


def test_store_skill_requirement_guides_assignment(tmp_path):
    database = tmp_path / "skill_requirement.sqlite"
    db.init_db(database)
    db.upsert_employee({"employee_id": "E001", "name": "レジ担当", "active": True,
                        "can_cashier": True, "max_consecutive_days": 5,
                        "min_work_days": 0, "max_work_days": 22}, database)
    db.upsert_employee({"employee_id": "E002", "name": "一般担当", "active": True,
                        "can_cashier": False, "max_consecutive_days": 5,
                        "min_work_days": 0, "max_work_days": 22}, database)
    db.replace_requirements("2026-08", [{"date": "2026-08-01", "shift_code": "D", "required_count": 1}], database)
    settings = db.get_store_settings(database)
    settings.update({"restaurant_mode": True, "require_english": False,
                     "require_new_product": False, "require_allergy": False})
    db.save_store_settings(settings, database)
    skill_rows = db.get_store_skill_requirements(database)
    cashier = next(row for row in skill_rows if row["skill_code"] == "cashier")
    cashier.update({"required_count": 1, "priority": "hard"})
    db.save_store_skill_requirements(skill_rows, database)

    result = generate_schedule("2026-08", 10, database)

    assert result["status"] == "success"
    assignment = next(item for item in result["assignments"] if item["date"] == "2026-08-01" and item["shift_code"] == "D")
    assert assignment["employee_id"] == "E001"
