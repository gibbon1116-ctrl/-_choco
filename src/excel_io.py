from __future__ import annotations

from datetime import date

from io import BytesIO
import math
from pathlib import Path
from typing import BinaryIO
import unicodedata

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import db
from .calendar_utils import display_date, is_weekend, month_dates
from .reports import employee_summary, request_violations, schedule_matrix, shift_summary
from .restaurant import RELATION_LABELS, ROLE_LABELS, restaurant_condition_checks
from .validators import validate_employee

ROOT = Path(__file__).resolve().parents[1]
EMPLOYEE_COLUMNS = ["employee_id", "name", "role", "skills", "active", "night_allowed",
                    "max_consecutive_days", "min_work_days", "max_work_days", "note",
                    "english_level", "can_cashier", "can_open", "can_close",
                    "can_handle_complaints", "can_explain_allergy", "is_new_staff",
                    "can_train_new_staff", "product_skill_ice", "product_skill_chocolate",
                    "product_skill_cookie", "new_product_skill", "can_manage_cash",
                    "can_hygiene_check", "peak_support_level"]
STAFF_SKILL_COLUMNS = ["employee_id", "name", "english_level", "can_cashier", "can_open",
                       "can_close", "can_handle_complaints", "can_explain_allergy", "is_new_staff",
                       "can_train_new_staff", "product_skill_ice", "product_skill_chocolate",
                       "product_skill_cookie", "new_product_skill", "can_manage_cash",
                       "can_hygiene_check", "peak_support_level"]
STAFF_RELATION_COLUMNS = ["employee_id_1", "employee_id_2", "relation_type", "priority", "weight", "active", "note"]
PRODUCT_CAMPAIGN_COLUMNS = ["product_name", "category", "start_date", "end_date",
                            "required_skill_level", "require_leader_first_week", "note"]
ROLE_REQUIREMENT_COLUMNS = ["target_month", "date", "shift_code", "role_code", "required_count", "priority"]
REQUIREMENT_COLUMNS = ["target_month", "date", "shift_code", "required_count"]
REQUEST_COLUMNS = ["target_month", "employee_id", "date", "request_type", "shift_code", "priority", "note"]


def _display_width(value: object) -> int:
    return sum(2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1
               for char in str(value or ""))


def _bool(value) -> bool:
    if isinstance(value, str):
        return value.strip().lower() in {"true", "1", "yes", "y", "はい", "可"}
    return bool(value)


def _read_excel(source: str | Path | BinaryIO) -> pd.DataFrame:
    return pd.read_excel(source, engine="openpyxl").fillna("")


def _require_columns(frame: pd.DataFrame, required: list[str]) -> None:
    missing = [column for column in required if column not in frame.columns]
    if missing:
        raise ValueError("Excelに必要な列がありません: " + ", ".join(missing))


def import_employees(source, db_path: str | Path | None = None) -> int:
    frame = _read_excel(source)
    _require_columns(frame, ["employee_id", "name"])
    prepared = []
    for _, row in frame.iterrows():
        data = {column: row.get(column, "") for column in EMPLOYEE_COLUMNS}
        data["active"] = _bool(row.get("active", True))
        data["night_allowed"] = _bool(row.get("night_allowed", True))
        for column in ("can_cashier", "can_open", "can_close", "can_handle_complaints",
                       "can_explain_allergy", "is_new_staff", "can_train_new_staff",
                       "can_manage_cash", "can_hygiene_check"):
            data[column] = _bool(row.get(column, False))
        data["english_level"] = str(row.get("english_level", "none") or "none").strip().lower()
        for column in ("product_skill_ice", "product_skill_chocolate", "product_skill_cookie",
                       "new_product_skill", "peak_support_level"):
            data[column] = int(row.get(column, 0) or 0)
        data["max_consecutive_days"] = int(row.get("max_consecutive_days", 5) or 5)
        data["min_work_days"] = int(row.get("min_work_days", 0) or 0)
        data["max_work_days"] = int(row.get("max_work_days", 31) or 31)
        errors = validate_employee(data)
        if errors:
            raise ValueError(f"職員ID {data['employee_id'] or '(空欄)'}: " + " ".join(errors))
        prepared.append(data)
    for data in prepared:
        db.upsert_employee(data, db_path)
    return len(frame)


def import_staff_skills(source, db_path: str | Path | None = None) -> int:
    """Update only restaurant skill fields without overwriting base employee data."""
    frame = _read_excel(source)
    _require_columns(frame, STAFF_SKILL_COLUMNS)
    existing = {e["employee_id"]: e for e in db.fetch_all("employees", db_path)}
    prepared = []
    for _, row in frame.iterrows():
        employee_id = str(row["employee_id"]).strip()
        if employee_id not in existing:
            raise ValueError(f"職員ID {employee_id} は職員マスタにありません。")
        data = dict(existing[employee_id])
        for column in STAFF_SKILL_COLUMNS:
            if column not in {"employee_id", "name"}:
                data[column] = row.get(column, data.get(column, 0))
        for column in ("can_cashier", "can_open", "can_close", "can_handle_complaints",
                       "can_explain_allergy", "is_new_staff", "can_train_new_staff",
                       "can_manage_cash", "can_hygiene_check"):
            data[column] = _bool(data.get(column, False))
        for column in ("product_skill_ice", "product_skill_chocolate", "product_skill_cookie",
                       "new_product_skill", "peak_support_level"):
            data[column] = int(data.get(column, 0) or 0)
        errors = validate_employee(data)
        if errors:
            raise ValueError(f"{employee_id}: " + " / ".join(errors))
        prepared.append(data)
    for item in prepared:
        db.upsert_employee(item, db_path)
    return len(prepared)


def import_requirements(source, target_month: str | None = None, db_path: str | Path | None = None) -> int:
    frame = _read_excel(source)
    _require_columns(frame, ["date", "shift_code", "required_count"])
    if target_month is None:
        if "target_month" not in frame.columns or frame.empty:
            raise ValueError("対象年月を指定してください。")
        target_month = str(frame.iloc[0]["target_month"])
    rows = [{"date": pd.to_datetime(row["date"]).date().isoformat(),
             "shift_code": str(row["shift_code"]).strip().upper(),
             "required_count": int(row["required_count"])} for _, row in frame.iterrows()]
    valid_dates = {day.isoformat() for day in month_dates(target_month)}
    shift_codes = {s["shift_code"] for s in db.fetch_all("shift_types", db_path) if s["is_work"]}
    for row in rows:
        if row["date"] not in valid_dates:
            raise ValueError(f"{row['date']} は対象年月 {target_month} の日付ではありません。")
        if row["shift_code"] not in shift_codes:
            raise ValueError(f"勤務区分 {row['shift_code']} は勤務区分マスタにありません。")
        if row["required_count"] < 0:
            raise ValueError("必要人数は0以上にしてください。")
    db.replace_requirements(target_month, rows, db_path)
    return len(rows)


def import_requests(source, target_month: str | None = None, db_path: str | Path | None = None) -> int:
    frame = _read_excel(source)
    _require_columns(frame, ["employee_id", "date", "request_type", "priority"])
    if target_month is None:
        target_month = str(frame.iloc[0].get("target_month", "")) if not frame.empty else ""
    valid_dates = {day.isoformat() for day in month_dates(target_month)}
    employee_ids = {e["employee_id"] for e in db.fetch_all("employees", db_path)}
    shift_codes = {s["shift_code"] for s in db.fetch_all("shift_types", db_path)}
    prepared = []
    for _, row in frame.iterrows():
        item = {"target_month": target_month, "employee_id": str(row["employee_id"]),
                "date": pd.to_datetime(row["date"]).date().isoformat(),
                "request_type": str(row["request_type"]).strip(),
                "shift_code": str(row.get("shift_code", "") or "").strip().upper(),
                "priority": str(row.get("priority", "soft")).strip(),
                "note": str(row.get("note", "") or "")}
        if item["employee_id"] not in employee_ids:
            raise ValueError(f"職員ID {item['employee_id']} は職員マスタにありません。")
        if item["date"] not in valid_dates:
            raise ValueError(f"{item['date']} は対象年月 {target_month} の日付ではありません。")
        if item["request_type"] not in {"off", "avoid", "prefer", "fixed"}:
            raise ValueError(f"希望種別 {item['request_type']} は使用できません。")
        if item["priority"] not in {"hard", "soft"}:
            raise ValueError("priority は hard または soft にしてください。")
        if item["request_type"] == "off":
            item["shift_code"] = "O"
        if item["shift_code"] not in shift_codes:
            raise ValueError(f"勤務区分 {item['shift_code']} は勤務区分マスタにありません。")
        prepared.append(item)
    with db.connect(db_path) as conn:
        conn.execute("DELETE FROM requests WHERE target_month=?", (target_month,))
    for item in prepared:
        db.add_request(item, db_path)
    return len(frame)


def import_staff_relations(source, db_path: str | Path | None = None) -> int:
    frame = _read_excel(source)
    _require_columns(frame, ["employee_id_1", "employee_id_2", "relation_type", "priority"])
    employee_ids = {e["employee_id"] for e in db.fetch_all("employees", db_path)}
    allowed = set(RELATION_LABELS)
    prepared = []
    for _, row in frame.iterrows():
        item = {"employee_id_1": str(row["employee_id_1"]), "employee_id_2": str(row["employee_id_2"]),
                "relation_type": str(row["relation_type"]), "priority": str(row.get("priority", "soft")),
                "weight": int(row.get("weight", 50) or 50), "active": _bool(row.get("active", True)),
                "note": str(row.get("note", "") or "")}
        if item["employee_id_1"] not in employee_ids or item["employee_id_2"] not in employee_ids:
            raise ValueError("スタッフ配置条件に職員マスタ未登録のIDがあります。")
        if item["employee_id_1"] == item["employee_id_2"]:
            raise ValueError("同じスタッフ同士は登録できません。")
        if item["relation_type"] not in allowed or item["priority"] not in {"hard", "soft"}:
            raise ValueError("配置ルールまたはpriorityの値を確認してください。")
        prepared.append(item)
    with db.connect(db_path) as conn:
        conn.execute("DELETE FROM staff_relations")
    for item in prepared:
        db.upsert_staff_relation(item, db_path)
    return len(prepared)


def import_product_campaigns(source, db_path: str | Path | None = None) -> int:
    frame = _read_excel(source)
    _require_columns(frame, ["product_name", "category", "start_date", "end_date", "required_skill_level"])
    prepared = []
    for _, row in frame.iterrows():
        item = {"product_name": str(row["product_name"]), "category": str(row["category"]),
                "start_date": pd.to_datetime(row["start_date"]).date().isoformat(),
                "end_date": pd.to_datetime(row["end_date"]).date().isoformat(),
                "required_skill_level": int(row.get("required_skill_level", 2) or 2),
                "require_leader_first_week": _bool(row.get("require_leader_first_week", True)),
                "note": str(row.get("note", "") or "")}
        if item["category"] not in {"ice", "chocolate", "cookie", "other"}:
            raise ValueError("category は ice / chocolate / cookie / other にしてください。")
        if item["start_date"] > item["end_date"] or item["required_skill_level"] not in {1, 2, 3}:
            raise ValueError(f"{item['product_name']} の期間または必要スキルレベルを確認してください。")
        prepared.append(item)
    with db.connect(db_path) as conn:
        conn.execute("DELETE FROM product_campaigns")
    for item in prepared:
        db.upsert_product_campaign(item, db_path)
    return len(prepared)


def import_role_requirements(source, target_month: str | None = None,
                             db_path: str | Path | None = None) -> int:
    frame = _read_excel(source)
    _require_columns(frame, ["date", "shift_code", "role_code", "required_count", "priority"])
    if target_month is None:
        target_month = str(frame.iloc[0].get("target_month", "")) if not frame.empty else ""
    valid_dates = {d.isoformat() for d in month_dates(target_month)}
    shift_codes = {s["shift_code"] for s in db.fetch_all("shift_types", db_path) if s["is_work"]}
    prepared = []
    for _, row in frame.iterrows():
        item = {"date": pd.to_datetime(row["date"]).date().isoformat(),
                "shift_code": str(row["shift_code"]).strip().upper(),
                "role_code": str(row["role_code"]).strip(),
                "required_count": int(row["required_count"]), "priority": str(row["priority"]).strip()}
        if item["date"] not in valid_dates or item["shift_code"] not in shift_codes:
            raise ValueError("役割別必要人数に対象年月外の日付または未登録の勤務区分があります。")
        if item["role_code"] not in ROLE_LABELS or item["priority"] not in {"hard", "soft"}:
            raise ValueError("role_code または priority の値を確認してください。")
        prepared.append(item)
    db.replace_role_requirements(target_month, prepared, db_path)
    return len(prepared)


def export_staff_relations_bytes(db_path=None) -> bytes:
    rows = db.fetch_all("staff_relations", db_path)
    return frame_to_excel_bytes(pd.DataFrame(rows, columns=["id"] + STAFF_RELATION_COLUMNS).drop(columns=["id"], errors="ignore"), "スタッフ配置条件")


def export_product_campaigns_bytes(db_path=None) -> bytes:
    rows = db.fetch_all("product_campaigns", db_path)
    return frame_to_excel_bytes(pd.DataFrame(rows, columns=["id"] + PRODUCT_CAMPAIGN_COLUMNS).drop(columns=["id"], errors="ignore"), "新商品キャンペーン")


def export_role_requirements_bytes(target_month: str, db_path=None) -> bytes:
    rows = db.fetch_all("role_requirements", db_path, where="target_month=?", params=(target_month,))
    return frame_to_excel_bytes(pd.DataFrame(rows, columns=["id"] + ROLE_REQUIREMENT_COLUMNS).drop(columns=["id"], errors="ignore"), "役割別必要人数")

def frame_to_excel_bytes(frame: pd.DataFrame, sheet_name: str) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.book[sheet_name]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1649C6")
        for column in worksheet.columns:
            width = min(max(_display_width(cell.value) for cell in column) + 2, 35)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = width
    return buffer.getvalue()


def export_employees_bytes(db_path=None) -> bytes:
    return frame_to_excel_bytes(pd.DataFrame(db.fetch_all("employees", db_path), columns=EMPLOYEE_COLUMNS), "職員マスタ")


def export_staff_skills_bytes(db_path=None) -> bytes:
    rows = db.fetch_all("employees", db_path)
    return frame_to_excel_bytes(pd.DataFrame(rows, columns=EMPLOYEE_COLUMNS)[STAFF_SKILL_COLUMNS], "スタッフスキル")


def export_requirements_bytes(target_month: str, db_path=None) -> bytes:
    rows = db.fetch_all("requirements", db_path, where="target_month=?", params=(target_month,))
    return frame_to_excel_bytes(pd.DataFrame(rows).drop(columns=["id"], errors="ignore"), "必要人数")


def export_requests_bytes(target_month: str, db_path=None) -> bytes:
    rows = db.fetch_all("requests", db_path, where="target_month=?", params=(target_month,))
    return frame_to_excel_bytes(pd.DataFrame(rows).drop(columns=["id"], errors="ignore"), "希望休・勤務希望")


def _style_sheet(ws) -> None:
    navy, blue, light = "10233F", "1649C6", "E7EBF2"
    thin = Side(style="thin", color="D8DEE8")
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for column in ws.columns:
        width = min(max(_display_width(cell.value) for cell in column) + 3, 40)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width
    for row_number, row in enumerate(ws.iter_rows(min_row=2), start=2):
        lines = 1
        for cell in row:
            column_width = ws.column_dimensions[get_column_letter(cell.column)].width or 10
            lines = max(lines, math.ceil(_display_width(cell.value) / max(8, column_width - 2)))
        if lines > 1:
            ws.row_dimensions[row_number].height = min(90, 16 * lines)
    ws.row_dimensions[1].height = 26


def export_schedule(schedule_id: int, output_dir: str | Path | None = None,
                    db_path: str | Path | None = None, *, admin_export: bool = False) -> Path:
    schedules = db.fetch_all("schedules", db_path, where="schedule_id=?", params=(schedule_id,))
    if not schedules:
        raise ValueError("出力対象の勤務表が見つかりません。")
    schedule = schedules[0]
    assignments = db.fetch_all("schedule_assignments", db_path, where="schedule_id=?", params=(schedule_id,))
    assignments = [{k: row[k] for k in ("employee_id", "date", "shift_code")} for row in assignments]
    shifts = db.fetch_all("shift_types", db_path)
    color_map = {s["shift_code"]: (s.get("color") or "FFFFFF").replace("#", "") for s in shifts}
    names = {e["employee_id"]: e["name"] for e in db.fetch_all("employees", db_path)}
    target_month = schedule["target_month"]

    wb = Workbook()
    wb.remove(wb.active)
    matrix = schedule_matrix(assignments, db_path)
    staff = employee_summary(assignments, db_path)
    shift_counts = shift_summary(assignments, db_path)
    violations = request_violations(target_month, assignments, db_path)

    ws = wb.create_sheet("勤務表")
    headers = ["職員名"] + [display_date(day) for day in matrix.columns] + ["勤務日数", "夜勤回数", "土日勤務"]
    ws.append(headers)
    for name, row in matrix.iterrows():
        stats = staff[staff["職員名"] == name].iloc[0]
        ws.append([name] + row.tolist() + [int(stats["勤務日数"]), int(stats["夜勤回数"]), int(stats["土日勤務"])])
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=1 + len(matrix.columns)):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=color_map.get(str(cell.value), "FFFFFF"))
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True, color="153E90" if cell.value != "O" else "5F6B7A")
    _style_sheet(ws)

    ws = wb.create_sheet("日別配置")
    ws.append(["日付", "曜日", "勤務区分", "配置職員"])
    for day in month_dates(target_month):
        day_text = day.isoformat()
        for shift in shifts:
            members = [names[a["employee_id"]] for a in assignments if a["date"] == day_text and a["shift_code"] == shift["shift_code"]]
            if members:
                ws.append([day_text, display_date(day).split("(")[-1].rstrip(")"),
                           f"{shift['shift_code']} {shift['shift_name']}", "、".join(members)])
    _style_sheet(ws)

    ws = wb.create_sheet("職員別集計")
    ws.append(staff.columns.tolist())
    for row in staff.itertuples(index=False, name=None): ws.append(list(row))
    _style_sheet(ws)

    ws = wb.create_sheet("勤務区分別集計")
    ws.append(shift_counts.columns.tolist())
    for row in shift_counts.itertuples(index=False, name=None): ws.append(list(row))
    _style_sheet(ws)

    ws = wb.create_sheet("希望休違反")
    violation_frame = pd.DataFrame(violations)
    if violation_frame.empty:
        ws.append(["結果"]); ws.append(["希望違反はありません。"])
    else:
        ws.append(violation_frame.columns.tolist())
        for row in violation_frame.itertuples(index=False, name=None): ws.append(list(row))
    _style_sheet(ws)

    ws = wb.create_sheet("ルール確認")
    ws.append(["項目", "内容"])
    rows = [("対象年月", target_month), ("作成日時", schedule["created_at"]), ("状態", schedule["status"]),
            ("ペナルティ合計", schedule["objective_value"]), ("計算時間（秒）", schedule["solver_wall_time"]),
            ("絶対条件", "1日1勤務・必要人数・夜勤可否・hard希望・最大連続勤務・夜勤明け休み"),
            ("最適化条件", "soft希望・勤務日数・勤務区分・夜勤回数・土日勤務の均等化、同点時のランダム分散")]
    for row in rows: ws.append(row)
    _style_sheet(ws)

    checks = restaurant_condition_checks(target_month, assignments, db_path)
    ws = wb.create_sheet("飲食店条件確認")
    if checks:
        check_frame = pd.DataFrame(checks)
        ws.append(check_frame.columns.tolist())
        for row in check_frame.itertuples(index=False, name=None):
            ws.append(list(row))
    else:
        ws.append(["結果"]); ws.append(["飲食店向け条件は無効、または確認対象がありません。"])
    _style_sheet(ws)

    ws = wb.create_sheet("スタッフスキル一覧")
    skill_frame = pd.DataFrame(db.fetch_all("employees", db_path), columns=EMPLOYEE_COLUMNS)[STAFF_SKILL_COLUMNS]
    ws.append(skill_frame.columns.tolist())
    for row in skill_frame.itertuples(index=False, name=None):
        ws.append(list(row))
    _style_sheet(ws)

    ws = wb.create_sheet("新商品イベント一覧")
    campaign_frame = pd.DataFrame(db.fetch_all("product_campaigns", db_path), columns=["id"] + PRODUCT_CAMPAIGN_COLUMNS)
    business_frame = pd.DataFrame(db.fetch_all("business_days", db_path, where="target_month=?", params=(target_month,)))
    ws.append(["区分", "名称", "カテゴリ", "開始日", "終了日", "需要レベル", "必要スキル", "備考"])
    for _, row in campaign_frame.iterrows():
        ws.append(["新商品", row["product_name"], row["category"], row["start_date"], row["end_date"],
                   "", row["required_skill_level"], row["note"]])
    if not business_frame.empty:
        for _, row in business_frame[business_frame["is_event_day"] == 1].iterrows():
            ws.append(["イベント", row["event_name"], "", row["date"], row["date"],
                       row["demand_level"], "", row["note"]])
    _style_sheet(ws)

    if admin_export:
        ws = wb.create_sheet("相性条件確認")
        ws.append(["スタッフ1", "スタッフ2", "配置ルール", "優先度", "重み", "結果", "管理者メモ"])
        relation_checks = [r for r in checks if r.get("確認項目") == "スタッフ配置条件" and r.get("結果") == "要確認"]
        for relation in db.fetch_all("staff_relations", db_path):
            token = f"{relation['employee_id_1']}・{relation['employee_id_2']}"
            violated = any(token in r.get("内容", "") for r in relation_checks)
            ws.append([names.get(relation["employee_id_1"], relation["employee_id_1"]),
                       names.get(relation["employee_id_2"], relation["employee_id_2"]),
                       RELATION_LABELS.get(relation["relation_type"], relation["relation_type"]),
                       relation["priority"], relation["weight"], "要確認" if violated else "充足",
                       relation.get("note", "")])
        _style_sheet(ws)
    output_dir = Path(output_dir or ROOT / "outputs")
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"勤務表_{target_month.replace('-', '')}_{schedule_id}_{'管理者確認用' if admin_export else '通常配布用'}.xlsx"
    wb.save(path)
    return path


def create_sample_files(target_month: str = "2026-08", sample_dir: str | Path | None = None) -> list[Path]:
    sample_dir = Path(sample_dir or ROOT / "sample_data")
    sample_dir.mkdir(parents=True, exist_ok=True)
    names = ["山田 太郎", "佐藤 花子", "鈴木 一郎", "高橋 美咲", "田中 裕子", "伊藤 大輔",
             "渡辺 恵美", "中村 健一", "小林 由香", "加藤 直樹", "吉田 真由美", "山本 翔平"]
    english = {1: "fluent", 2: "conversational", 3: "basic", 4: "conversational"}
    employees = []
    for i, name in enumerate(names, 1):
        employees.append({
            "employee_id": f"E{i:03d}", "name": name,
            "role": "店長" if i == 1 else ("リーダー" if i in {2, 3} else "スタッフ"),
            "skills": "飲食店接客", "active": True, "night_allowed": False,
            "max_consecutive_days": 5, "min_work_days": 10, "max_work_days": 22, "note": "",
            "english_level": english.get(i, "none"), "can_cashier": i <= 8,
            "can_open": i <= 6, "can_close": i in {1, 2, 4, 5, 6, 7},
            "can_handle_complaints": i in {1, 2, 4}, "can_explain_allergy": i in {1, 2, 4, 5},
            "is_new_staff": i in {11, 12}, "can_train_new_staff": i in {1, 2, 3},
            "product_skill_ice": 3 if i in {1, 3} else (2 if i <= 8 else 1),
            "product_skill_chocolate": 3 if i in {2, 3} else (2 if i <= 8 else 1),
            "product_skill_cookie": 3 if i in {1, 2} else (2 if i <= 8 else 1),
            "new_product_skill": 3 if i in {1, 2, 3} else (2 if i <= 7 else 1),
            "can_manage_cash": i in {1, 2, 4}, "can_hygiene_check": i in {1, 6},
            "peak_support_level": 3 if i in {1, 2} else (2 if i <= 7 else 1),
        })
    requirements = []
    role_requirements = []
    for day in month_dates(target_month):
        template = {"E": 2, "D": 4, "L": 2} if is_weekend(day) else {"E": 2, "D": 3, "L": 2}
        for code, count in template.items():
            requirements.append({"target_month": target_month, "date": day.isoformat(),
                                 "shift_code": code, "required_count": count})
        role_requirements.extend([
            {"target_month": target_month, "date": day.isoformat(), "shift_code": "E",
             "role_code": "opener", "required_count": 1, "priority": "hard"},
            {"target_month": target_month, "date": day.isoformat(), "shift_code": "L",
             "role_code": "closer", "required_count": 1, "priority": "hard"},
            {"target_month": target_month, "date": day.isoformat(), "shift_code": "D",
             "role_code": "cashier", "required_count": 1, "priority": "soft"},
        ])
    requests = [
        {"target_month": target_month, "employee_id": "E001", "date": f"{target_month}-05", "request_type": "off", "shift_code": "O", "priority": "hard", "note": "私用"},
        {"target_month": target_month, "employee_id": "E002", "date": f"{target_month}-10", "request_type": "off", "shift_code": "O", "priority": "soft", "note": "私用"},
        {"target_month": target_month, "employee_id": "E003", "date": f"{target_month}-12", "request_type": "avoid", "shift_code": "L", "priority": "soft", "note": ""},
        {"target_month": target_month, "employee_id": "E004", "date": f"{target_month}-03", "request_type": "fixed", "shift_code": "E", "priority": "hard", "note": "開店研修"},
        {"target_month": target_month, "employee_id": "E005", "date": f"{target_month}-18", "request_type": "prefer", "shift_code": "D", "priority": "soft", "note": ""},
        {"target_month": target_month, "employee_id": "E008", "date": f"{target_month}-22", "request_type": "off", "shift_code": "O", "priority": "hard", "note": "家族行事"},
    ]
    relations = [
        {"employee_id_1": "E001", "employee_id_2": "E011", "relation_type": "mentor_pair", "priority": "soft", "weight": 150, "active": True, "note": "新人フォロー"},
        {"employee_id_1": "E002", "employee_id_2": "E012", "relation_type": "mentor_pair", "priority": "soft", "weight": 150, "active": True, "note": "新人フォロー"},
        {"employee_id_1": "E005", "employee_id_2": "E006", "relation_type": "avoid_together", "priority": "soft", "weight": 180, "active": True, "note": "配置バランス"},
        {"employee_id_1": "E007", "employee_id_2": "E008", "relation_type": "prefer_peak_pair", "priority": "soft", "weight": 100, "active": True, "note": "繁忙対応"},
        {"employee_id_1": "E009", "employee_id_2": "E010", "relation_type": "never_together", "priority": "hard", "weight": 300, "active": True, "note": "同時配置禁止"},
    ]
    campaigns = [
        {"product_name": "季節のアイスサンド", "category": "ice", "start_date": f"{target_month}-05",
         "end_date": f"{target_month}-12", "required_skill_level": 2, "require_leader_first_week": True, "note": "夏季新商品"},
        {"product_name": "カカオクッキー", "category": "cookie", "start_date": f"{target_month}-20",
         "end_date": f"{target_month}-27", "required_skill_level": 2, "require_leader_first_week": True, "note": "重点販売"},
    ]
    files = {
        "employees": sample_dir / "sample_employees.xlsx",
        "skills": sample_dir / "staff_skills.xlsx",
        "requirements": sample_dir / "sample_requirements.xlsx",
        "requests": sample_dir / "sample_requests.xlsx",
        "relations": sample_dir / "staff_relations.xlsx",
        "campaigns": sample_dir / "product_campaigns.xlsx",
        "roles": sample_dir / "role_requirements.xlsx",
    }
    files["employees"].write_bytes(frame_to_excel_bytes(pd.DataFrame(employees)[EMPLOYEE_COLUMNS], "職員マスタ"))
    files["skills"].write_bytes(frame_to_excel_bytes(pd.DataFrame(employees)[STAFF_SKILL_COLUMNS], "スタッフスキル"))
    files["requirements"].write_bytes(frame_to_excel_bytes(pd.DataFrame(requirements), "必要人数"))
    files["requests"].write_bytes(frame_to_excel_bytes(pd.DataFrame(requests), "希望休・勤務希望"))
    files["relations"].write_bytes(frame_to_excel_bytes(pd.DataFrame(relations), "スタッフ配置条件"))
    files["campaigns"].write_bytes(frame_to_excel_bytes(pd.DataFrame(campaigns), "新商品キャンペーン"))
    files["roles"].write_bytes(frame_to_excel_bytes(pd.DataFrame(role_requirements), "役割別必要人数"))
    return list(files.values())


def load_sample_data(target_month: str = "2026-08", db_path: str | Path | None = None,
                     sample_dir: str | Path | None = None) -> dict:
    db.init_db(db_path)
    paths = create_sample_files(target_month, sample_dir)
    lookup = {path.name: path for path in paths}
    result = {
        "employees": import_employees(lookup["sample_employees.xlsx"], db_path),
        "requirements": import_requirements(lookup["sample_requirements.xlsx"], target_month, db_path),
        "requests": import_requests(lookup["sample_requests.xlsx"], target_month, db_path),
        "relations": import_staff_relations(lookup["staff_relations.xlsx"], db_path),
        "campaigns": import_product_campaigns(lookup["product_campaigns.xlsx"], db_path),
        "role_requirements": import_role_requirements(lookup["role_requirements.xlsx"], target_month, db_path),
    }
    settings = db.get_store_settings(db_path)
    settings.update({"store_name": "路面店A", "business_hours": "10:00-21:00",
                     "weekday_required": 7, "weekend_required": 8, "restaurant_mode": True,
                     "require_english": True, "english_priority": "hard", "required_english_count": 1,
                     "require_english_per_shift": False, "require_new_product": True,
                     "new_product_priority": "soft", "require_allergy": True, "allergy_priority": "soft"})
    db.save_store_settings(settings, db_path)
    for day_number, name, demand, new_product in [
        (5, "新商品発売", "very_high", True), (15, "商店街セール", "high", False),
        (22, "近隣イベント", "high", False),
    ]:
        day = date.fromisoformat(f"{target_month}-{day_number:02d}")
        db.upsert_business_day({"target_month": target_month, "date": day.isoformat(), "is_open": True,
                                "is_weekend": day.weekday() >= 5, "is_event_day": True, "event_name": name,
                                "demand_level": demand, "new_product_active": new_product, "note": "サンプル"}, db_path)
    result["events"] = 3
    return result
